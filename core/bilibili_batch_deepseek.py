# -*- coding: utf-8 -*-
"""
bilibili_batch_deepseek: Excel批量视频转Markdown工具

功能：
  - 读取Excel中的B站视频列表
  - 批量下载视频（yt-dlp）
  - 使用本地Whisper进行语音识别
  - 使用DeepSeek生成Markdown文档
  - 自动跳过已下载和已处理的视频

流程：
  Excel → 下载视频 → Whisper ASR → DeepSeek LLM → Markdown

依赖：
  pip install openai-whisper yt-dlp openpyxl requests

用法：
  python bilibili_batch_deepseek.py --excel 视频列表.xlsx
  python bilibili_batch_deepseek.py --excel 视频列表.xlsx --whisper-model small
  python bilibili_batch_deepseek.py --excel 视频列表.xlsx --limit 5
"""

__version__ = "1.0.0"

import os
import sys
import json
import time
import hashlib
import subprocess
from pathlib import Path
from typing import Optional
from datetime import datetime

# 添加项目路径
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

# ===================== 配置 =====================

# DeepSeek 配置
DEEPSEEK_API_KEY = "sk-d2e27860aa3243668a95a85e2bf8edcb"
DEEPSEEK_BASE_URL = "https://api.deepseek.com/v1"
DEEPSEEK_MODEL = "deepseek-chat"

# Whisper 配置
WHISPER_MODEL = "medium"
WHISPER_LANGUAGE = "zh"

# 输出目录
VIDEO_DIR = PROJECT_ROOT / "output" / "batch_videos"      # 视频下载目录
DOCS_DIR = PROJECT_ROOT / "output" / "batch_docs"         # Markdown输出目录
ASR_CACHE_DIR = PROJECT_ROOT / "output" / "batch_asr_cache"  # ASR缓存目录
PROGRESS_FILE = PROJECT_ROOT / "output" / "batch_progress.json"  # 进度记录

# yt-dlp 配置
YTDLP_TIMEOUT = 600  # 下载超时（秒）

# Cookies 文件路径
COOKIE_FILE = PROJECT_ROOT / "config" / "cookies.txt"


# ===================== Cookies 处理 =====================

def load_cookies() -> dict:
    """从 cookies.txt 加载 cookies（Netscape 格式）"""
    if not COOKIE_FILE.exists():
        return {}
    
    cookies = {}
    try:
        with open(COOKIE_FILE, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                parts = line.split("\t")
                if len(parts) >= 7:
                    name = parts[5]
                    value = parts[6]
                    cookies[name] = value
    except Exception as e:
        print(f"[警告] 读取 cookies 失败: {e}")
    
    return cookies


def cookie_dict_to_str(cookies: dict) -> str:
    """将 cookie 字典转为字符串"""
    return "; ".join([f"{k}={v}" for k, v in cookies.items()])


def _find_ffmpeg() -> str:
    """查找 ffmpeg 路径"""
    import shutil
    path = shutil.which("ffmpeg")
    if path:
        return path
    
    # 检查项目 bin 目录
    bin_dir = PROJECT_ROOT / "bin"
    if bin_dir.exists():
        ffmpeg_path = bin_dir / "ffmpeg.exe"
        if ffmpeg_path.exists():
            return str(ffmpeg_path)
    
    return ""


# ===================== 进度管理 =====================

def _load_progress() -> dict:
    """加载处理进度"""
    if PROGRESS_FILE.exists():
        try:
            with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            pass
    return {"downloaded": {}, "processed": {}}


def _save_progress(progress: dict):
    """保存处理进度"""
    PROGRESS_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


def _is_downloaded(progress: dict, bvid: str) -> bool:
    """检查视频是否已下载"""
    return bvid in progress.get("downloaded", {})


def _is_processed(progress: dict, bvid: str) -> bool:
    """检查视频是否已处理成Markdown"""
    return bvid in progress.get("processed", {})


def _mark_downloaded(progress: dict, bvid: str, video_path: str):
    """标记视频已下载"""
    if "downloaded" not in progress:
        progress["downloaded"] = {}
    progress["downloaded"][bvid] = {
        "path": video_path,
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    _save_progress(progress)


def _mark_processed(progress: dict, bvid: str, md_path: str):
    """标记视频已处理"""
    if "processed" not in progress:
        progress["processed"] = {}
    progress["processed"][bvid] = {
        "path": md_path,
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    _save_progress(progress)


# ===================== Excel 读取 =====================

def read_excel_videos(excel_path: str) -> list[dict]:
    """
    读取Excel中的视频列表
    返回: [{"bvid": "...", "title": "...", "url": "...", "category": "...", ...}, ...]
    """
    import openpyxl
    
    print(f"[读取] 打开Excel: {excel_path}")
    
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb.active
    
    # 读取表头
    headers = [cell.value for cell in ws[1]]
    print(f"[读取] 表头: {headers}")
    
    videos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        
        # 构建字典
        data = {}
        for i, header in enumerate(headers):
            if header and i < len(row):
                data[header] = row[i]
        
        # 必须有 bvid 或播放链接
        bvid = data.get("视频讯号(bvid)") or data.get("bvid") or ""
        url = data.get("播放链接") or data.get("url") or ""
        
        if not bvid and not url:
            continue
        
        # 从URL提取bvid
        if not bvid and url:
            if "/video/" in url:
                bvid = url.split("/video/")[-1].split("/")[0].split("?")[0]
        
        if not bvid:
            continue
        
        videos.append({
            "bvid": bvid,
            "title": str(data.get("视频标题") or data.get("title") or "").strip(),
            "url": url or f"https://www.bilibili.com/video/{bvid}",
            "category": str(data.get("类别") or "").strip(),
            "keyword": str(data.get("检索关键词") or "").strip(),
            "search_word": str(data.get("B站检索词") or "").strip(),
            "author": str(data.get("视频作者") or "").strip(),
            "duration": str(data.get("播放时长") or "").strip(),
        })
    
    wb.close()
    print(f"[读取] 共读取 {len(videos)} 个视频")
    return videos


# ===================== 视频下载 =====================

def _safe_filename(name: str) -> str:
    """生成安全的文件名"""
    # 移除或替换非法字符
    illegal = '<>:"/\\|?*'
    for c in illegal:
        name = name.replace(c, '_')
    # 限制长度
    if len(name) > 100:
        name = name[:100]
    return name.strip()


def download_video(url: str, bvid: str, title: str, output_dir: str) -> Optional[str]:
    """
    使用 yt-dlp 下载视频（与 bilibili_search_download_v2_ui.py 相同方式）
    返回: 视频文件路径，失败返回 None
    """
    import sys
    
    VIDEO_DIR.mkdir(parents=True, exist_ok=True)
    
    # 文件名：bvid_标题
    safe_title = _safe_filename(title) if title else bvid
    output_template = str(VIDEO_DIR / f"{bvid}_{safe_title}.%(ext)s")
    
    # 检查是否已存在
    for ext in ["mp4", "mkv", "webm", "flv"]:
        check_path = str(VIDEO_DIR / f"{bvid}_{safe_title}.{ext}")
        if os.path.exists(check_path):
            print(f"  [下载] 视频已存在: {check_path}")
            return check_path
    
    print(f"  [下载] 开始下载: {title[:50]}...")
    print(f"  [下载] URL: {url}")
    
    # yt-dlp 下载命令（与 bilibili_search_download_v2_ui.py 相同）
    cmd = [
        sys.executable, "-m", "yt_dlp",
        "--no-warnings",
        "-f", "bestvideo+bestaudio/best",
        "--merge-output-format", "mp4",
        "-o", output_template,
        "--no-playlist",
        "--socket-timeout", "30",
        "--retries", "3",
    ]
    
    # 添加 ffmpeg 路径
    ffmpeg_path = _find_ffmpeg()
    if ffmpeg_path:
        cmd.extend(["--ffmpeg-location", os.path.dirname(ffmpeg_path)])
    
    # 添加 cookies（使用 --add-header 方式，与 bilibili_search_download_v2_ui.py 相同）
    cookies = load_cookies()
    if cookies:
        cookie_str = cookie_dict_to_str(cookies)
        cmd.extend(["--add-header", f"Cookie:{cookie_str}"])
        print(f"  [下载] 使用cookies")
    
    cmd.append(url)
    
    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            encoding='utf-8',
            errors='ignore',
            timeout=YTDLP_TIMEOUT
        )
        
        if result.returncode != 0:
            print(f"  [下载] 失败: {result.stderr[:200]}")
            return None
        
        # 查找下载的文件
        for ext in ["mp4", "mkv", "webm", "flv"]:
            check_path = str(VIDEO_DIR / f"{bvid}_{safe_title}.{ext}")
            if os.path.exists(check_path):
                print(f"  [下载] 完成: {check_path}")
                return check_path
        
        # 尝试查找目录中的新文件
        for f in VIDEO_DIR.iterdir():
            if f.suffix.lower() in [".mp4", ".mkv", ".webm", ".flv"]:
                if bvid in f.stem or safe_title in f.stem:
                    print(f"  [下载] 完成: {f}")
                    return str(f)
        
        print(f"  [下载] 失败: 找不到下载的文件")
        return None
        
    except subprocess.TimeoutExpired:
        print(f"  [下载] 超时")
        return None
    except Exception as e:
        print(f"  [下载] 错误: {e}")
        return None


# ===================== Whisper ASR =====================

def _asr_cache_path(file_path: str) -> str:
    """生成ASR缓存路径"""
    file_hash = hashlib.md5(os.path.abspath(file_path).encode('utf-8')).hexdigest()[:16]
    return os.path.join(ASR_CACHE_DIR, f"{file_hash}.whisper.json")


def _load_asr_cache(file_path: str) -> Optional[list]:
    """加载ASR缓存"""
    cache_file = _asr_cache_path(file_path)
    if not os.path.isfile(cache_file):
        return None
    
    try:
        with open(cache_file, "r", encoding="utf-8") as f:
            cache = json.load(f)
    except:
        return None
    
    # 验证文件是否变化
    stat = os.stat(file_path)
    if cache.get("video_size") != stat.st_size:
        return None
    
    return cache.get("sentences", [])


def _save_asr_cache(file_path: str, sentences: list):
    """保存ASR缓存"""
    ASR_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache_file = _asr_cache_path(file_path)
    stat = os.stat(file_path)
    
    cache = {
        "video_path": os.path.abspath(file_path),
        "video_size": stat.st_size,
        "transcribe_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "sentences": sentences,
    }
    
    with open(cache_file, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


def transcribe_with_whisper(video_path: str) -> list[dict]:
    """使用Whisper进行语音识别"""
    # 检查缓存
    cached = _load_asr_cache(video_path)
    if cached:
        print(f"  [ASR] 命中缓存，共 {len(cached)} 句")
        return cached
    
    print(f"  [ASR] 开始转写: {Path(video_path).name}")
    print(f"  [ASR] 模型: {WHISPER_MODEL}")
    
    # 提取音频
    audio_path = video_path + ".audio.wav"
    print(f"  [ASR] 提取音频...")
    
    cmd = [
        "ffmpeg", "-i", video_path,
        "-vn", "-acodec", "pcm_s16le",
        "-ar", "16000", "-ac", "1",
        "-y", audio_path
    ]
    
    result = subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        encoding='utf-8',
        errors='ignore'
    )
    
    if result.returncode != 0:
        raise RuntimeError(f"ffmpeg 提取音频失败: {result.stderr}")
    
    try:
        import whisper
        
        print(f"  [ASR] 加载模型...")
        model = whisper.load_model(WHISPER_MODEL)
        
        print(f"  [ASR] 开始识别...")
        result = model.transcribe(
            audio_path,
            language=WHISPER_LANGUAGE,
            verbose=False
        )
        
        sentences = []
        for seg in result.get("segments", []):
            sentences.append({
                "text": seg["text"].strip(),
                "start": seg.get("start", 0.0),
                "end": seg.get("end", 0.0),
            })
        
        print(f"  [ASR] 完成: 共 {len(sentences)} 句")
        
        # 保存缓存
        _save_asr_cache(video_path, sentences)
        
        return sentences
        
    finally:
        if os.path.exists(audio_path):
            os.remove(audio_path)


# ===================== DeepSeek LLM =====================

def _call_deepseek(prompt: str) -> str:
    """调用DeepSeek API"""
    import requests
    
    headers = {
        "Authorization": f"Bearer {DEEPSEEK_API_KEY}",
        "Content-Type": "application/json",
    }
    
    payload = {
        "model": DEEPSEEK_MODEL,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.7,
        "max_tokens": 4096,
    }
    
    print(f"  [DeepSeek] 调用 {DEEPSEEK_MODEL}...")
    
    response = requests.post(
        f"{DEEPSEEK_BASE_URL}/chat/completions",
        headers=headers,
        json=payload,
        timeout=120
    )
    
    if response.status_code != 200:
        raise RuntimeError(f"DeepSeek API 失败: {response.status_code}")
    
    return response.json()["choices"][0]["message"]["content"].strip()


def _extract_json(content: str) -> dict:
    """提取JSON"""
    if content.startswith("```"):
        lines = content.split("\n")
        lines = [l for l in lines if not l.strip().startswith("```")]
        content = "\n".join(lines)
    return json.loads(content)


def clean_with_deepseek(sentences: list[dict]) -> list[dict]:
    """使用DeepSeek清洗文本"""
    numbered = [f"[{i}] {s['text']}" for i, s in enumerate(sentences)]
    full_text = "\n".join(numbered)
    
    prompt = f"""请对以下视频转写文本进行清洗：
1. 润色语句，使其更通顺
2. 去除无关内容（求关注、广告、口头禅等）
3. 保留所有有效内容

输出JSON格式：{{"cleaned": ["句子1", "句子2", ...]}}

转写文本：
{full_text}"""
    
    content = _call_deepseek(prompt)
    result = _extract_json(content)
    
    cleaned = [{"text": t} for t in result.get("cleaned", [])]
    print(f"  [DeepSeek] 清洗: {len(sentences)} → {len(cleaned)} 句")
    return cleaned


def structure_with_deepseek(sentences: list[dict]) -> dict:
    """使用DeepSeek结构化处理"""
    numbered = [f"[{i}] {s['text']}" for i, s in enumerate(sentences)]
    full_text = "\n".join(numbered)
    
    prompt = f"""请分析以下视频转写文本，输出主题、概要和分段：

要求：
1. 主题：一句话概括
2. 概要：2-3句话
3. 分段：按语义分段，每段给标题

输出JSON：{{"topic": "主题", "summary": "概要", "segments": [{{"title": "标题", "start": 0, "end": 5}}]}}

转写文本：
{full_text}"""
    
    content = _call_deepseek(prompt)
    result = _extract_json(content)
    
    # 修正索引范围
    last_idx = len(sentences) - 1
    for seg in result.get("segments", []):
        seg["end"] = min(seg.get("end", last_idx), last_idx)
        seg["start"] = min(seg.get("start", 0), last_idx)
    
    print(f"  [DeepSeek] 主题: {result.get('topic', '未知')}")
    print(f"  [DeepSeek] 分段: {len(result.get('segments', []))} 段")
    return result


# ===================== Markdown 生成 =====================

def generate_markdown(sentences: list[dict], structure: dict, video_info: dict) -> str:
    """生成Markdown文档"""
    lines = [
        f"# {structure.get('topic', '视频内容')}",
        "",
        f"> **视频**: {video_info.get('title', '未知')}",
        f"> **作者**: {video_info.get('author', '未知')}",
        f"> **类别**: {video_info.get('category', '未分类')}",
        f"> **生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "## 概要",
        "",
        structure.get("summary", ""),
        "",
    ]
    
    for seg in structure.get("segments", []):
        lines.append(f"## {seg.get('title', '段落')}")
        lines.append("")
        
        start = seg.get("start", 0)
        end = seg.get("end", len(sentences) - 1)
        
        for i in range(start, end + 1):
            if i < len(sentences):
                lines.append(sentences[i]["text"])
                lines.append("")
    
    return "\n".join(lines)


# ===================== 主流程 =====================

def process_single_video(video: dict, progress: dict) -> Optional[str]:
    """处理单个视频"""
    bvid = video["bvid"]
    title = video.get("title", bvid)
    url = video.get("url", f"https://www.bilibili.com/video/{bvid}")
    
    print(f"\n{'='*60}")
    print(f"处理: {title[:50]}...")
    print(f"BVID: {bvid}")
    print(f"{'='*60}")
    
    # 检查是否已处理
    if _is_processed(progress, bvid):
        md_path = progress["processed"][bvid]["path"]
        if os.path.exists(md_path):
            print(f"[跳过] 已处理: {md_path}")
            return md_path
    
    # 1. 下载视频
    video_path = None
    if _is_downloaded(progress, bvid):
        video_path = progress["downloaded"][bvid]["path"]
        if os.path.exists(video_path):
            print(f"[下载] 已存在: {video_path}")
        else:
            video_path = None
    
    if not video_path:
        video_path = download_video(url, bvid, title, str(VIDEO_DIR))
        if not video_path:
            print("[错误] 下载失败")
            return None
        _mark_downloaded(progress, bvid, video_path)
    
    # 2. Whisper ASR
    try:
        sentences = transcribe_with_whisper(video_path)
    except Exception as e:
        print(f"[错误] ASR失败: {e}")
        return None
    
    if not sentences:
        print("[错误] ASR结果为空")
        return None
    
    # 3. DeepSeek 清洗
    try:
        cleaned = clean_with_deepseek(sentences)
    except Exception as e:
        print(f"[错误] 清洗失败: {e}")
        return None
    
    # 4. DeepSeek 结构化
    try:
        structure = structure_with_deepseek(cleaned)
    except Exception as e:
        print(f"[错误] 结构化失败: {e}")
        return None
    
    # 5. 生成 Markdown
    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    md_filename = f"{bvid}_{_safe_filename(title[:50])}.md"
    md_path = str(DOCS_DIR / md_filename)
    
    markdown = generate_markdown(cleaned, structure, video)
    
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(markdown)
    
    _mark_processed(progress, bvid, md_path)
    
    print(f"\n[完成] 输出: {md_path}")
    return md_path


def main():
    import argparse
    
    global WHISPER_MODEL
    
    parser = argparse.ArgumentParser(description="Excel批量视频转Markdown")
    parser.add_argument("--excel", "-e", required=True, help="Excel文件路径")
    parser.add_argument("--limit", "-l", type=int, default=0, help="限制处理数量（0=全部）")
    parser.add_argument("--whisper-model", "-m", default=WHISPER_MODEL, 
                       help=f"Whisper模型 (默认: {WHISPER_MODEL})")
    parser.add_argument("--reset", action="store_true", help="重置进度记录")
    args = parser.parse_args()
    
    WHISPER_MODEL = args.whisper_model
    
    # 检查Excel
    if not os.path.exists(args.excel):
        print(f"[错误] Excel不存在: {args.excel}")
        sys.exit(1)
    
    # 加载进度
    if args.reset:
        progress = {"downloaded": {}, "processed": {}}
        _save_progress(progress)
        print("[进度] 已重置")
    else:
        progress = _load_progress()
    
    # 读取视频列表
    videos = read_excel_videos(args.excel)
    if not videos:
        print("[错误] Excel中没有有效视频")
        sys.exit(1)
    
    # 限制数量
    if args.limit > 0:
        videos = videos[:args.limit]
    
    # 统计
    total = len(videos)
    skipped = sum(1 for v in videos if _is_processed(progress, v["bvid"]))
    pending = total - skipped
    
    print(f"\n{'='*60}")
    print(f"批量处理统计")
    print(f"{'='*60}")
    print(f"总计: {total} 个视频")
    print(f"已处理: {skipped} 个")
    print(f"待处理: {pending} 个")
    print(f"Whisper模型: {WHISPER_MODEL}")
    print(f"{'='*60}\n")
    
    if pending == 0:
        print("[完成] 所有视频已处理")
        return
    
    # 处理视频
    success = 0
    failed = 0
    
    for i, video in enumerate(videos, 1):
        if _is_processed(progress, video["bvid"]):
            continue
        
        print(f"\n[{i}/{total}] ", end="")
        
        result = process_single_video(video, progress)
        if result:
            success += 1
        else:
            failed += 1
    
    # 汇总
    print(f"\n{'='*60}")
    print(f"处理完成")
    print(f"{'='*60}")
    print(f"成功: {success} 个")
    print(f"失败: {failed} 个")
    print(f"视频目录: {VIDEO_DIR}")
    print(f"文档目录: {DOCS_DIR}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
