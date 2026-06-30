# -*- coding: utf-8 -*-
"""
bilibili-search-down-v3: B站视频搜索汇总工具

功能：
  - 读取Excel中的检索词和类别
  - 调用B站搜索API获取视频列表信息
  - 使用大模型对搜索结果进行相关性重排序
  - 输出为Excel（仅搜索，不下载视频）

输入Excel格式（至少包含以下列）：
  - 类别（第一列）：如"商机"、"教程"等
  - 检索关键词（第二列）：如"Python教程"、"AI入门"等
  - B站检索词（第三列）：用于B站搜索的关键词

输出Excel格式：
  类别 | 检索关键词 | B站检索词 | 视频讯号(bvid) | 视频标题 | 播放链接 |
  视频作者 | 播放时长 | 弹幕数量 | 收藏数量 | 点赞数量 | 播放次数 | 发布时间

用法：
  python bilibili_search_down_v3.py
  python bilibili_search_down_v3.py --input 商机组.xlsx --output 结果.xlsx

依赖：
  pip install requests openpyxl dashscope

注意：
  - 需要登录B站获取cookie以提高搜索配额
  - 请在 config/cookies.txt 放置cookies.txt文件
  - 需要设置 DASHSCOPE_API_KEY 环境变量用于大模型排序
"""

__version__ = "3.1.0"

import hashlib
import json
import os
import sys
import time
import urllib.parse
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
import requests

# 添加项目路径
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

# 导入大模型配置
try:
    from parser.config import DASHSCOPE_API_KEY, LLM_MODEL
except ImportError:
    DASHSCOPE_API_KEY = os.getenv("DASHSCOPE_API_KEY", "")
    LLM_MODEL = "qwen-plus"

# 大模型相关配置
LLM_TOP_N = 5  # 大模型排序后保留前N条

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/125.0.0.0 Safari/537.36"
    ),
    "Referer": "https://www.bilibili.com",
    "Accept": "application/json, text/plain, */*",
    "Origin": "https://search.bilibili.com",
}

# WBI 签名字符表（去除特殊字符）
_ILLEGAL_CHARS = "!'()*"


def _mixin_key(orig: str) -> str:
    """Mixin key 用于 WBI 签名"""
    mixin_key_tab = [
        46, 47, 18, 2, 53, 8, 23, 32, 15, 50, 10, 31, 58, 3, 45, 35, 27, 43, 5, 49,
        33, 9, 42, 19, 29, 28, 14, 39, 12, 38, 41, 13, 4, 48, 24, 40, 1, 5, 15, 44,
        28, 16, 6, 51, 21, 17, 0, 56, 25, 0, 8, 4, 9, 57, 43, 24, 30, 55, 46, 57,
        33, 51, 37, 46, 47, 0, 58, 43, 32, 24, 41, 54, 49, 26, 21, 27, 56, 14, 52,
        22, 18, 32, 57, 12, 29, 3, 43, 47, 37, 38, 20, 48, 51, 4, 25, 56, 22, 29,
        50, 10, 36, 15, 21, 5, 56, 58, 51, 24, 17, 43, 47, 35, 17, 0, 6, 36, 53,
        14, 5, 41, 25, 26, 51, 52, 58, 43, 26, 18, 16, 43, 57, 43, 4, 49, 30, 41,
    ]
    return bytes(c ^ i for c, i in zip(orig.encode(), mixin_key_tab[:len(orig)])).decode()


def _extract_key(url: str) -> str:
    """从 URL 中提取 key"""
    match = [x for x in url.split("/") if x and "=" in x]
    if not match:
        return ""
    params = urllib.parse.parse_qs(match[0])
    return params["params"][0]


def _find_cookie_file() -> str:
    """按优先级搜索 cookies.txt"""
    script_dir = Path(__file__).resolve().parent
    candidates = [
        script_dir / "cookies.txt",
        script_dir.parent / "config" / "cookies.txt",
        script_dir.parent / "cookies.txt",
    ]
    for p in candidates:
        if p.is_file():
            return str(p)
    return str(candidates[0])


COOKIE_FILE = _find_cookie_file()


def load_cookies() -> dict:
    """加载 cookies"""
    if not os.path.exists(COOKIE_FILE):
        print(f"[提示] 未找到cookies文件: {COOKIE_FILE}")
        print("       将使用未登录模式，搜索配额可能受限")
        return {}

    cookie_str = open(COOKIE_FILE, encoding="utf-8").read().strip()
    if not cookie_str:
        return {}

    cookies = {}
    for line in cookie_str.split("\n"):
        line = line.strip()
        if line and not line.startswith("#"):
            parts = line.split("\t")
            if len(parts) >= 7:
                cookies[parts[5]] = parts[6]
    return cookies


def _fetch_wbi_keys(session: requests.Session) -> tuple[str, str]:
    """获取 WBI 签名密钥"""
    resp = session.get(
        "https://api.bilibili.com/x/web-interface/nav",
        headers=HEADERS, timeout=10,
    )
    resp.raise_for_status()
    data = resp.json()
    if data.get("code") != 0:
        raise RuntimeError(f"WBI密钥获取失败: {data.get('message')}")
    wbi = data["data"]["wbi_img"]
    return _extract_key(wbi["img_url"]), _extract_key(wbi["sub_url"])


def sign_params(params: dict, img_key: str, sub_key: str) -> dict:
    """对参数进行 WBI 签名"""
    key = _mixin_key(img_key + sub_key)
    params = {**params, "wts": int(time.time())}
    cleaned = {}
    for k, v in sorted(params.items()):
        cleaned[k] = str(v).translate(str.maketrans("", "", _ILLEGAL_CHARS))
    query = urllib.parse.urlencode(cleaned)
    cleaned["w_rid"] = hashlib.md5((query + key).encode()).hexdigest()
    return cleaned


# ===================== Excel 读取 =====================

def read_input_excel(filepath: str) -> list[dict]:
    """
    读取输入Excel
    格式：类别 | 检索关键词 | B站检索词
    只用 B站检索词 作为搜索条件
    返回: [{"category": "商机", "search_keyword": "AI教程", "bilibili_keyword": "AI教程"}, ...]
    """
    print(f"[读取] 打开Excel: {filepath}")

    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    tasks = []
    header_row = None
    
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if not row or not any(row):
            continue
        
        # 第一行可能是表头，检查并记录列位置
        if i == 0:
            row_str = [str(c).strip() if c else "" for c in row]
            if "类别" in row_str or "检索关键词" in row_str or "B站检索词" in row_str:
                header_row = row_str
                continue
        
        # 解析列数据
        category = ""
        search_keyword = ""
        bilibili_keyword = ""
        
        if header_row:
            # 根据表头定位列
            try:
                cat_idx = header_row.index("类别")
                category = str(row[cat_idx]).strip() if row[cat_idx] else ""
            except (ValueError, IndexError):
                pass
            try:
                sk_idx = header_row.index("检索关键词")
                search_keyword = str(row[sk_idx]).strip() if row[sk_idx] else ""
            except (ValueError, IndexError):
                pass
            try:
                bk_idx = header_row.index("B站检索词")
                bilibili_keyword = str(row[bk_idx]).strip() if row[bk_idx] else ""
            except (ValueError, IndexError):
                pass
        else:
            # 没有表头，按默认顺序：类别 | 检索关键词 | B站检索词
            if len(row) >= 3:
                category = str(row[0]).strip() if row[0] else ""
                search_keyword = str(row[1]).strip() if row[1] else ""
                bilibili_keyword = str(row[2]).strip() if row[2] else ""
            elif len(row) >= 2:
                category = str(row[0]).strip() if row[0] else ""
                bilibili_keyword = str(row[1]).strip() if row[1] else ""
            elif len(row) >= 1:
                bilibili_keyword = str(row[0]).strip() if row[0] else ""
        
        # 必须有B站检索词才加入任务
        if not bilibili_keyword:
            continue
        
        # 跳过标题行
        if bilibili_keyword in ["B站检索词", "检索词", "搜索词"]:
            continue
        
        tasks.append({
            "category": category or "未分类",
            "search_keyword": search_keyword or bilibili_keyword,  # 检索关键词默认为B站检索词
            "bilibili_keyword": bilibili_keyword,
        })

    wb.close()
    print(f"[读取] 共读取 {len(tasks)} 个检索词")
    return tasks


# ===================== B站搜索 =====================

def search_videos(keyword: str, page_size: int = 30) -> list[dict]:
    """
    搜索B站视频（仅获取信息，不下载）
    过滤：弹幕数量为0、无播放链接、播放时长为空的视频会被忽略
    """
    session = requests.Session()
    cookies = load_cookies()
    if cookies:
        session.cookies = requests.utils.cookiejar_from_dict(cookies)

    try:
        img_key, sub_key = _fetch_wbi_keys(session)
    except Exception as e:
        print(f"[警告] WBI密钥获取失败({e})，搜索可能受限")
        img_key = sub_key = ""

    params = {
        "search_type": "video",
        "keyword": keyword,
        "page": 1,
        "page_size": page_size,
        "order": "totalrank",  # 综合排序
    }

    if img_key and sub_key:
        params = sign_params(params, img_key, sub_key)

    try:
        resp = session.get(
            "https://api.bilibili.com/x/web-interface/wbi/search/type",
            params=params, headers=HEADERS, timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        print(f"[错误] 请求失败: {e}")
        return []

    if data.get("code") != 0:
        print(f"[错误] API返回: code={data.get('code')}, message={data.get('message')}")
        return []

    results = data.get("data", {}).get("result", [])
    if not results:
        print(f"[搜索] 关键词「{keyword}」无结果")
        return []

    videos = []
    filtered_count = 0
    
    for item in results:
        # 清理标题中的HTML标签
        title = item.get("title", "").replace('<em class="keyword">', "").replace("</em>", "")
        bvid = item.get("bvid", "")
        url = f"https://www.bilibili.com/video/{bvid}" if bvid else ""

        # 格式化发布时间
        pubdate = item.get("pubdate", 0)
        if pubdate:
            pubdate_str = datetime.fromtimestamp(pubdate).strftime("%Y-%m-%d %H:%M:%S")
        else:
            pubdate_str = ""

        # 格式化时长
        duration = item.get("duration", "")
        
        # 获取弹幕数量
        danmaku = item.get("video_review", 0) or 0

        # ===== 过滤逻辑 =====
        # 1. 没有播放链接
        if not url or not bvid:
            filtered_count += 1
            continue
        # 2. 播放时长为空
        if not duration:
            filtered_count += 1
            continue
        # 3. 弹幕数量为0
        if danmaku <= 0:
            filtered_count += 1
            continue

        videos.append({
            "bvid": bvid,
            "title": title,
            "url": url,
            "author": item.get("author", ""),
            "duration": duration,
            "danmaku": danmaku,
            "favorites": item.get("favorites", 0),
            "like": item.get("like", 0),
            "play": item.get("play", 0),
            "pubdate": pubdate_str,
        })

    if filtered_count > 0:
        print(f"[过滤] 忽略 {filtered_count} 条无效视频（无链接/无时长/无弹幕）")

    return videos


# ===================== 大模型重排序 =====================

def _call_llm(prompt: str) -> str:
    """调用大模型并返回原始文本内容"""
    api_key = DASHSCOPE_API_KEY
    if not api_key:
        raise ValueError("请设置 DASHSCOPE_API_KEY 环境变量")
    
    import dashscope
    from dashscope import Generation
    
    dashscope.api_key = api_key
    
    response = Generation.call(
        model=LLM_MODEL,
        messages=[{"role": "user", "content": prompt}],
        result_format="message",
    )
    
    if response.status_code != 200:
        raise RuntimeError(f"LLM调用失败: {response.message}")
    
    return response.output.choices[0].message.content.strip()


def _extract_json(content: str) -> list:
    """从LLM返回中提取JSON列表"""
    if content.startswith("```"):
        lines = content.split("\n")
        lines = [l for l in lines if not l.strip().startswith("```")]
        content = "\n".join(lines)
    try:
        result = json.loads(content)
        if isinstance(result, list):
            return result
        elif isinstance(result, dict) and "ranked" in result:
            return result["ranked"]
        else:
            return []
    except json.JSONDecodeError as e:
        raise RuntimeError(f"LLM返回的JSON解析失败: {e}\n原始内容: {content[:500]}")


def rerank_videos_by_llm(
    videos: list[dict],
    category: str,
    search_keyword: str,
    bilibili_keyword: str,
    top_n: int = LLM_TOP_N
) -> list[dict]:
    """
    使用大模型对视频进行相关性重排序
    
    Args:
        videos: 视频列表
        category: 类别
        search_keyword: 检索关键词
        bilibili_keyword: B站检索词
        top_n: 保留前N条
    
    Returns:
        重排序后的视频列表（最多top_n条）
    """
    if not videos:
        return []
    
    if len(videos) <= top_n:
        return videos
    
    # 构造带编号的视频标题列表
    numbered_titles = []
    for i, video in enumerate(videos):
        numbered_titles.append(f"[{i}] {video['title']}")
    titles_text = "\n".join(numbered_titles)
    
    prompt = f"""你是一个专业的视频内容相关性评估专家。请根据以下三个条件，对视频标题进行相关性排序：

**评估条件：**
1. 类别：{category}
2. 检索关键词：{search_keyword}
3. B站检索词：{bilibili_keyword}

**排序要求：**
- 根据视频标题与上述三个条件的**综合相关性**进行排序
- 优先保留与"类别"和"检索关键词"高度相关的视频
- 考虑标题是否准确反映了搜索意图
- 忽略标题党、低质量、明显不相关的视频

**输出格式：**
请严格按以下JSON格式输出排序结果（索引从0开始），不要输出其他内容：
{{
  "ranked": [0, 3, 1, 5, 2]
}}

其中 ranked 数组包含排序后的视频索引，按相关性从高到低排列。

**待排序的视频标题列表：**
{titles_text}

请输出相关性最高的前 {top_n} 个视频的索引。"""

    try:
        print(f"  [LLM] 调用大模型重排序 {len(videos)} 条视频...")
        content = _call_llm(prompt)
        ranked_indices = _extract_json(content)
        
        # 根据排序结果重新组织视频列表
        ranked_videos = []
        for idx in ranked_indices[:top_n]:
            if 0 <= idx < len(videos):
                ranked_videos.append(videos[idx])
        
        print(f"  [LLM] 重排序完成，保留 {len(ranked_videos)} 条最相关视频")
        return ranked_videos
        
    except Exception as e:
        print(f"  [LLM] 重排序失败: {e}，使用原始顺序")
        return videos[:top_n]


# ===================== Excel 输出 =====================

def write_output_excel(filepath: str, data: list[dict]):
    """
    写入结果Excel
    列：类别 | 检索关键词 | B站检索词 | 视频讯号(bvid) | 视频标题 | 播放链接 |
        视频作者 | 播放时长 | 弹幕数量 | 收藏数量 | 点赞数量 | 播放次数 | 发布时间
    """
    print(f"[写入] 生成Excel: {filepath}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "B站视频搜索结果"

    # 表头
    headers = [
        "类别", "检索关键词", "B站检索词", "视频讯号(bvid)", "视频标题", "播放链接",
        "视频作者", "播放时长", "弹幕数量", "收藏数量", "点赞数量", "播放次数", "发布时间"
    ]
    ws.append(headers)

    # 数据
    for item in data:
        ws.append([
            item.get("category", ""),
            item.get("search_keyword", ""),   # 检索关键词
            item.get("bilibili_keyword", ""), # B站检索词
            item.get("bvid", ""),
            item.get("title", ""),
            item.get("url", ""),
            item.get("author", ""),
            item.get("duration", ""),
            item.get("danmaku", ""),
            item.get("favorites", ""),
            item.get("like", ""),
            item.get("play", ""),
            item.get("pubdate", ""),
        ])

    # 调整列宽
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)

    wb.save(filepath)
    print(f"[完成] 共写入 {len(data)} 条记录")


# ===================== 主程序 =====================

def main():
    import argparse

    parser = argparse.ArgumentParser(description="B站视频搜索汇总工具 v3")
    parser.add_argument("--input", "-i", default=None, help="输入Excel路径")
    parser.add_argument("--output", "-o", default=None, help="输出Excel路径")
    parser.add_argument("--count", "-c", type=int, default=30, help="每个关键词搜索数量(默认30)")
    args = parser.parse_args()

    # 自动查找输入文件
    if not args.input:
        # 查找项目根目录下的xlsx文件
        for f in PROJECT_ROOT.glob("*.xlsx"):
            if "商机" in f.name:
                args.input = str(f)
                break
        if not args.input:
            for f in PROJECT_ROOT.glob("*.xlsx"):
                args.input = str(f)
                break

    if not args.input:
        print("[错误] 请指定输入Excel文件: --input <path>")
        sys.exit(1)

    if not os.path.exists(args.input):
        print(f"[错误] 输入文件不存在: {args.input}")
        sys.exit(1)

    # 默认输出文件名
    if not args.output:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        args.output = str(PROJECT_ROOT / f"bilibili_search_result_{timestamp}.xlsx")

    # 读取任务列表
    tasks = read_input_excel(args.input)
    if not tasks:
        print("[错误] Excel中没有找到有效的检索词")
        sys.exit(1)

    # 搜索并汇总
    print(f"\n{'='*50}")
    print(f"开始搜索，共 {len(tasks)} 个关键词，每个获取 {args.count} 条结果")
    print(f"大模型重排序后保留前 {LLM_TOP_N} 条最相关视频")
    print(f"{'='*50}\n")

    all_results = []
    total_videos = 0

    for i, task in enumerate(tasks, 1):
        category = task["category"]
        search_keyword = task["search_keyword"]      # 检索关键词（用于输出）
        bilibili_keyword = task["bilibili_keyword"]  # B站检索词（用于搜索）

        print(f"[{i}/{len(tasks)}] 搜索: 「{bilibili_keyword}」(类别: {category})")

        # 1. 搜索视频
        videos = search_videos(bilibili_keyword, args.count)
        
        if not videos:
            print(f"       → 无有效视频\n")
            continue
        
        print(f"       → 搜索到 {len(videos)} 条有效视频")
        
        # 2. 大模型重排序
        if DASHSCOPE_API_KEY and len(videos) > LLM_TOP_N:
            videos = rerank_videos_by_llm(
                videos, category, search_keyword, bilibili_keyword, LLM_TOP_N
            )
        else:
            if not DASHSCOPE_API_KEY:
                print(f"  [提示] 未设置 DASHSCOPE_API_KEY，跳过大模型重排序")
            videos = videos[:LLM_TOP_N]

        # 3. 添加到结果
        for video in videos:
            video["category"] = category
            video["search_keyword"] = search_keyword
            video["bilibili_keyword"] = bilibili_keyword
            all_results.append(video)

        total_videos += len(videos)
        print(f"       → 最终保留 {len(videos)} 条视频\n")

        # 避免请求过快
        if i < len(tasks):
            time.sleep(0.5)

    # 写入结果
    print(f"{'='*50}")
    print(f"搜索完成！共 {len(tasks)} 个关键词，{total_videos} 条最终视频")
    print(f"{'='*50}")

    write_output_excel(args.output, all_results)

    print(f"\n结果已保存到: {args.output}")


if __name__ == "__main__":
    main()
