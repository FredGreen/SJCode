# -*- coding: utf-8 -*-
"""
SJCode 视频内容处理工作台 - 主窗口

功能：
  - Excel 任务上传和管理
  - B站视频下载
  - 视频列表展示和 ASR 队列管理
  - 视频转 Markdown（ASR + LLM）
  - 商机提炼总结
  - SQLite 数据库存储所有数据
"""

import os
import sys
import subprocess
from pathlib import Path
from datetime import datetime

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QTableWidget, QTableWidgetItem, QCheckBox,
    QFileDialog, QMessageBox, QGroupBox, QHeaderView, QAbstractItemView,
    QStatusBar, QListWidget, QStackedWidget, QComboBox, QProgressBar,
)
from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtGui import QColor

# 确保项目根目录在 sys.path 中
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from config.settings import (
    OUTPUT_DIR, VIDEO, TASKS, HISTORY, DOCS, SUMMARY, ASR_CACHE
)
import database


# ===================== 下载线程 =====================

class DownloadWorker(QThread):
    """视频下载工作线程"""
    progress = Signal(str)
    finished = Signal(bool, str)
    video_added = Signal(dict)
    progress_value = Signal(int, int)

    def __init__(self, tasks: list[dict], parent=None):
        super().__init__(parent)
        self.tasks = tasks
        self._running = True

    def run(self):
        print(f"[DownloadWorker] 线程启动，共 {len(self.tasks)} 个任务")
        try:
            from core.bilibili_search_download_v2_ui import download_keyword_videos

            for i, task in enumerate(self.tasks):
                if not self._running:
                    break

                task_id = task.get('id')
                keyword = task["keyword"]
                order = task.get("order_type", "totalrank")
                limit = task.get("limit_count", 5)

                print(f"[DownloadWorker] 开始任务 [{i+1}/{len(self.tasks)}]: {keyword}")
                self.progress.emit(f"[{i+1}/{len(self.tasks)}] 下载: {keyword}")
                self.progress_value.emit(i + 1, len(self.tasks))

                # 更新任务状态为运行中
                if task_id:
                    database.update_task_status(task_id, 'running')

                videos = download_keyword_videos(
                    keyword, order=order, limit=limit,
                    progress_callback=lambda msg: self.progress.emit(f"  {msg}")
                )

                print(f"[DownloadWorker] 下载完成，找到 {len(videos)} 个视频")

                # 保存视频到数据库
                for video in videos:
                    video_id = database.add_video(
                        task_id=task_id,
                        bvid=video.get('bvid', ''),
                        title=video.get('title', ''),
                        file_path=video.get('path', ''),
                        file_size=video.get('size', 0),
                        duration=video.get('duration', ''),
                        favorites=0,
                        likes=0,
                        plays=0,
                        keyword=keyword
                    )
                    video['db_id'] = video_id
                    self.video_added.emit(video)

                # 更新任务状态为完成
                if task_id:
                    database.update_task_status(task_id, 'completed')

                self.progress.emit(f"  完成: {keyword} ({len(videos)} 个视频)")

            print("[DownloadWorker] 所有任务完成")
            self.finished.emit(True, "所有任务已完成")

        except Exception as e:
            print(f"[DownloadWorker] 异常: {str(e)}")
            import traceback
            traceback.print_exc()
            self.finished.emit(False, f"下载失败: {str(e)}")

    def stop(self):
        self._running = False


# ===================== ASR 转文字线程 =====================

class ASRWorker(QThread):
    """视频转文字工作线程"""
    progress = Signal(str)
    finished = Signal(bool, str)
    transcription_added = Signal(dict)

    def __init__(self, tasks: list[dict], parent=None):
        super().__init__(parent)
        self.tasks = tasks

    def run(self):
        print(f"[ASRWorker] 线程启动，共 {len(self.tasks)} 个任务")
        BASE_DIR = Path(__file__).resolve().parent.parent

        try:
            for i, task in enumerate(self.tasks):
                transcription_id = task.get('id')
                video_path = task.get('video_path', '')

                if not video_path or not Path(video_path).exists():
                    print(f"[ASRWorker] 文件不存在: {video_path}")
                    database.update_transcription(transcription_id, 'failed', error_message='文件不存在')
                    continue

                print(f"[ASRWorker] 处理 [{i+1}/{len(self.tasks)}]: {Path(video_path).name}")
                self.progress.emit(f"[{i+1}/{len(self.tasks)}] 转文字: {Path(video_path).name[:30]}...")

                database.update_transcription(transcription_id, 'running')

                cmd = [
                    sys.executable, "-m", "parser.main", video_path,
                    "--source", "B站"
                ]

                result = subprocess.run(
                    cmd,
                    capture_output=True, text=True,
                    encoding='utf-8', errors='ignore',
                    cwd=str(BASE_DIR),
                    timeout=600
                )

                print(f"[ASRWorker] 返回码: {result.returncode}")

                if result.returncode == 0:
                    # 解析输出文件路径
                    video_name = Path(video_path).stem
                    output_path = str(DOCS / f"{video_name}.md")

                    database.update_transcription(transcription_id, 'completed', output_path=output_path)

                    self.transcription_added.emit({
                        'id': transcription_id,
                        'video_path': video_path,
                        'output_path': output_path,
                        'status': 'completed'
                    })
                    self.progress.emit(f"  完成: {Path(video_path).name[:30]}")
                else:
                    error_msg = result.stderr or '未知错误'
                    database.update_transcription(transcription_id, 'failed', error_message=error_msg)
                    self.progress.emit(f"  失败: {error_msg[:50]}")

            print("[ASRWorker] 所有任务完成")
            self.finished.emit(True, "所有转文字任务已完成")

        except Exception as e:
            print(f"[ASRWorker] 异常: {str(e)}")
            import traceback
            traceback.print_exc()
            self.finished.emit(False, f"转文字失败: {str(e)}")


# ===================== 提炼总结线程 =====================

class SummaryWorker(QThread):
    """提炼总结工作线程"""
    progress = Signal(str)
    finished = Signal(bool, str)
    summary_added = Signal(dict)

    def __init__(self, tasks: list[dict], parent=None):
        super().__init__(parent)
        self.tasks = tasks

    def run(self):
        print(f"[SummaryWorker] 线程启动，共 {len(self.tasks)} 个任务")
        BASE_DIR = Path(__file__).resolve().parent.parent

        try:
            from parser.summary import summarize_file

            for i, task in enumerate(self.tasks):
                summary_id = task.get('id')
                input_path = task.get('input_path', '')

                if not input_path or not Path(input_path).exists():
                    print(f"[SummaryWorker] 文件不存在: {input_path}")
                    database.update_summary(summary_id, 'failed', error_message='文件不存在')
                    continue

                print(f"[SummaryWorker] 处理 [{i+1}/{len(self.tasks)}]: {Path(input_path).name}")
                self.progress.emit(f"[{i+1}/{len(self.tasks)}] 提炼: {Path(input_path).name[:30]}...")

                database.update_summary(summary_id, 'running')

                output_path = summarize_file(input_path)

                if output_path and Path(output_path).exists():
                    database.update_summary(summary_id, 'completed', output_path=output_path)
                    self.summary_added.emit({
                        'id': summary_id,
                        'input_path': input_path,
                        'output_path': output_path,
                        'status': 'completed'
                    })
                    self.progress.emit(f"  完成: {Path(input_path).name[:30]}")
                else:
                    database.update_summary(summary_id, 'failed', error_message='输出文件未生成')
                    self.progress.emit(f"  失败: 输出文件未生成")

            print("[SummaryWorker] 所有任务完成")
            self.finished.emit(True, "所有提炼总结已完成")

        except Exception as e:
            print(f"[SummaryWorker] 异常: {str(e)}")
            import traceback
            traceback.print_exc()
            self.finished.emit(False, f"提炼失败: {str(e)}")


# ===================== 主窗口 =====================

class VideoProcessorApp(QMainWindow):
    def __init__(self):
        super().__init__()

        # 初始化数据库
        database.init_database()

        self.setWindowTitle("SJCode 视频内容处理工作台")
        self.setMinimumSize(1300, 900)

        # 数据存储
        self.preview_data = []
        self.current_tasks = []

        # 工作线程
        self.download_worker = None
        self.asr_worker = None
        self.summary_worker = None

        self.setup_ui()
        self.refresh_all_tables()

        self.setStyleSheet("""
            QMainWindow { background-color: #f5f7fa; }
            QGroupBox { font-weight: bold; border: 1px solid #d0d7de; border-radius: 6px; margin-top: 10px; padding-top: 10px; background-color: white; }
            QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 5px 0 5px; color: #24292f; }
            QPushButton { background-color: #2c974b; color: white; border: none; border-radius: 6px; padding: 6px 12px; font-weight: bold; }
            QPushButton:hover { background-color: #2c6e3f; }
            QPushButton:pressed { background-color: #1b4d2a; }
            QPushButton#secondary { background-color: #6c757d; }
            QPushButton#secondary:hover { background-color: #5a6268; }
            QPushButton#dangerBtn { background-color: #d73a49; }
            QPushButton#dangerBtn:hover { background-color: #b31d28; }
            QTableWidget { border: 1px solid #e1e4e8; border-radius: 4px; alternate-background-color: #f8f9fa; gridline-color: #e1e4e8; }
            QHeaderView::section { background-color: #f6f8fa; padding: 4px; border: none; border-right: 1px solid #e1e4e8; border-bottom: 1px solid #e1e4e8; font-weight: bold; }
            QComboBox, QLineEdit { padding: 4px; border: 1px solid #d0d7de; border-radius: 4px; }
            QListWidget { background-color: white; border: none; border-right: 1px solid #e1e4e8; outline: none; }
            QListWidget::item { padding: 12px 8px; border-bottom: 1px solid #e1e4e8; }
            QListWidget::item:selected { background-color: #e2e6ea; color: #24292f; }
            QListWidget::item:hover { background-color: #f0f2f4; }
            QProgressBar { border: 1px solid #d0d7de; border-radius: 4px; text-align: center; }
            QProgressBar::chunk { background-color: #2c974b; border-radius: 3px; }
        """)

    # ========== 界面布局 ==========

    def setup_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # 左侧菜单
        self.sidebar = QListWidget()
        self.sidebar.setMaximumWidth(180)
        self.sidebar.setMinimumWidth(150)
        for item in ["📋 任务配置", "🎬 视频库", "📝 转文字任务", "✨ 提炼总结", "⚙️ 设置"]:
            self.sidebar.addItem(item)
        self.sidebar.currentRowChanged.connect(self.switch_page)
        main_layout.addWidget(self.sidebar)

        # 右侧堆叠区域
        self.stacked_widget = QStackedWidget()
        main_layout.addWidget(self.stacked_widget, 1)

        self.page_task_config = self.create_task_config_page()
        self.page_video_library = self.create_video_library_page()
        self.page_transcribe = self.create_transcribe_page()
        self.page_summary = self.create_summary_page()
        self.page_settings = self.create_settings_page()

        self.stacked_widget.addWidget(self.page_task_config)
        self.stacked_widget.addWidget(self.page_video_library)
        self.stacked_widget.addWidget(self.page_transcribe)
        self.stacked_widget.addWidget(self.page_summary)
        self.stacked_widget.addWidget(self.page_settings)

        self.statusBar = QStatusBar()
        self.setStatusBar(self.statusBar)
        self.statusBar.showMessage("就绪 | SQLite 数据库已连接")

    def switch_page(self, index):
        self.stacked_widget.setCurrentIndex(index)

    # ========== 页面1: 任务配置 ==========

    def create_task_config_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setSpacing(12)

        # 上传区域
        upload_group = QGroupBox("📂 上传Excel文件")
        upload_layout = QHBoxLayout(upload_group)
        self.upload_btn = QPushButton("选择文件")
        self.upload_btn.clicked.connect(self.upload_excel)
        self.file_label = QLabel("未选择文件")
        upload_layout.addWidget(self.upload_btn)
        upload_layout.addWidget(self.file_label)
        upload_layout.addStretch()
        layout.addWidget(upload_group)

        # 预览表格
        preview_group = QGroupBox("📄 Excel数据预览")
        preview_layout = QVBoxLayout(preview_group)
        self.preview_table = QTableWidget()
        self.preview_table.setColumnCount(5)
        self.preview_table.setHorizontalHeaderLabels(["选择", "检索关键词", "排序", "TOP-N", "渠道"])
        self.preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.preview_table.setAlternatingRowColors(True)
        self.preview_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        preview_layout.addWidget(self.preview_table)

        btn_layout = QHBoxLayout()
        self.add_selected_btn = QPushButton("✅ 将选中行加入任务")
        self.add_selected_btn.clicked.connect(self.add_selected_to_tasks)
        self.add_all_btn = QPushButton("📌 全部加入任务")
        self.add_all_btn.clicked.connect(self.add_all_to_tasks)
        btn_layout.addWidget(self.add_selected_btn)
        btn_layout.addWidget(self.add_all_btn)
        btn_layout.addStretch()
        preview_layout.addLayout(btn_layout)
        layout.addWidget(preview_group)

        # 待执行任务表格
        task_group = QGroupBox("⏳ 待执行任务列表")
        task_layout = QVBoxLayout(task_group)
        self.task_table = QTableWidget()
        self.task_table.setColumnCount(5)
        self.task_table.setHorizontalHeaderLabels(["关键词", "排序方式", "数量", "渠道", "状态"])
        self.task_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.task_table.setAlternatingRowColors(True)
        task_layout.addWidget(self.task_table)

        action_layout = QHBoxLayout()
        self.clear_pending_btn = QPushButton("🗑️ 清空已完成任务")
        self.clear_pending_btn.setObjectName("dangerBtn")
        self.clear_pending_btn.clicked.connect(self.clear_completed_tasks)

        self.start_crawl_btn = QPushButton("🚀 开启视频抓取任务")
        self.start_crawl_btn.clicked.connect(self.start_crawling)

        self.progress_bar = QProgressBar()

        action_layout.addWidget(self.clear_pending_btn)
        action_layout.addWidget(self.start_crawl_btn)
        action_layout.addStretch()
        task_layout.addLayout(action_layout)
        task_layout.addWidget(self.progress_bar)
        layout.addWidget(task_group)

        return page

    # ========== 页面2: 视频库 ==========

    def create_video_library_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        video_group = QGroupBox("🎬 历史下载视频库")
        video_layout = QVBoxLayout(video_group)
        self.history_table = QTableWidget()
        self.history_table.setColumnCount(6)
        self.history_table.setHorizontalHeaderLabels(
            ["视频名称", "关键词", "大小(MB)", "时长", "收藏数", "操作"]
        )
        self.history_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.history_table.setAlternatingRowColors(True)
        video_layout.addWidget(self.history_table)

        btn_layout = QHBoxLayout()
        self.refresh_video_btn = QPushButton("🔄 刷新列表")
        self.refresh_video_btn.setObjectName("secondary")
        self.refresh_video_btn.clicked.connect(self.refresh_video_table)

        self.add_to_asr_btn = QPushButton("🎤 将选中视频加入转文字")
        self.add_to_asr_btn.clicked.connect(self.add_selected_videos_to_asr)

        btn_layout.addWidget(self.refresh_video_btn)
        btn_layout.addWidget(self.add_to_asr_btn)
        btn_layout.addStretch()
        video_layout.addLayout(btn_layout)
        layout.addWidget(video_group)
        return page

    # ========== 页面3: 转文字任务 ==========

    def create_transcribe_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        asr_group = QGroupBox("🎙️ 视频转文字任务队列")
        asr_layout = QVBoxLayout(asr_group)
        self.asr_table = QTableWidget()
        self.asr_table.setColumnCount(4)
        self.asr_table.setHorizontalHeaderLabels(["视频名称", "状态", "输出文件", "操作"])
        self.asr_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        asr_layout.addWidget(self.asr_table)

        btn_layout = QHBoxLayout()
        self.refresh_asr_btn = QPushButton("🔄 刷新列表")
        self.refresh_asr_btn.setObjectName("secondary")
        self.refresh_asr_btn.clicked.connect(self.refresh_asr_table)

        self.start_transcribe_btn = QPushButton("🔄 开启视频转文章任务")
        self.start_transcribe_btn.clicked.connect(self.start_transcription)

        btn_layout.addWidget(self.refresh_asr_btn)
        btn_layout.addWidget(self.start_transcribe_btn)
        btn_layout.addStretch()
        asr_layout.addLayout(btn_layout)
        layout.addWidget(asr_group)

        # 已完成的转文字列表
        trans_group = QGroupBox("📄 已生成的转文字文件")
        trans_layout = QVBoxLayout(trans_group)
        self.transcription_table = QTableWidget()
        self.transcription_table.setColumnCount(3)
        self.transcription_table.setHorizontalHeaderLabels(["文件名", "源视频", "时间"])
        self.transcription_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        trans_layout.addWidget(self.transcription_table)
        layout.addWidget(trans_group)
        return page

    # ========== 页面4: 提炼总结 ==========

    def create_summary_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        summary_group = QGroupBox("✨ AI提炼总结")
        summary_layout = QVBoxLayout(summary_group)
        self.summary_table = QTableWidget()
        self.summary_table.setColumnCount(4)
        self.summary_table.setHorizontalHeaderLabels(["总结文件名", "源文件", "状态", "操作"])
        self.summary_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        summary_layout.addWidget(self.summary_table)

        btn_layout = QHBoxLayout()
        self.refresh_summary_btn = QPushButton("🔄 刷新列表")
        self.refresh_summary_btn.setObjectName("secondary")
        self.refresh_summary_btn.clicked.connect(self.refresh_summary_table)

        self.generate_summary_btn = QPushButton("🧠 从转文字文件生成总结")
        self.generate_summary_btn.clicked.connect(self.generate_summary_from_trans)

        btn_layout.addWidget(self.refresh_summary_btn)
        btn_layout.addWidget(self.generate_summary_btn)
        btn_layout.addStretch()
        summary_layout.addLayout(btn_layout)
        layout.addWidget(summary_group)
        return page

    # ========== Excel处理 ==========

    def upload_excel(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择Excel文件", "", "Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            self.file_label.setText(file_path)
            self.load_excel_preview(file_path)

    def load_excel_preview(self, file_path):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))

            if len(rows) < 2:
                QMessageBox.warning(self, "警告", "Excel至少需要标题行和数据行")
                return

            self.preview_data = []
            for row in rows[1:]:
                if len(row) >= 3 and row[0]:
                    keyword = str(row[0]).strip()
                    sort_type = str(row[1]).strip() if len(row) > 1 and row[1] else "totalrank"
                    topn = str(row[2]).strip() if len(row) > 2 and row[2] else "5"
                    if keyword:
                        # 检查是否重复
                        is_duplicate = database.check_keyword_exists(keyword)
                        self.preview_data.append({
                            "keyword": keyword,
                            "sort": sort_type,
                            "count": topn,
                            "duplicate": is_duplicate
                        })

            self.preview_table.setRowCount(len(self.preview_data))
            for idx, data in enumerate(self.preview_data):
                chk = QCheckBox()
                self.preview_table.setCellWidget(idx, 0, chk)
                self.preview_table.setItem(idx, 1, QTableWidgetItem(data["keyword"]))
                self.preview_table.setItem(idx, 2, QTableWidgetItem(data["sort"]))
                self.preview_table.setItem(idx, 3, QTableWidgetItem(data["count"]))

                # 渠道下拉框
                channel_combo = QComboBox()
                channel_combo.addItems(["B站", "抖音", "youtube"])
                self.preview_table.setCellWidget(idx, 4, channel_combo)

                # 如果是重复关键词，标红
                if data.get("duplicate"):
                    for col in range(1, 4):
                        item = self.preview_table.item(idx, col)
                        if item:
                            item.setForeground(QColor("red"))

            self.statusBar.showMessage(f"已加载 {len(self.preview_data)} 行数据（红色为重复关键词）")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"读取Excel失败: {str(e)}")

    def add_selected_to_tasks(self):
        selected_rows = []
        for row in range(self.preview_table.rowCount()):
            chk_widget = self.preview_table.cellWidget(row, 0)
            if chk_widget and chk_widget.isChecked():
                selected_rows.append(row)

        if not selected_rows:
            QMessageBox.information(self, "提示", "请先勾选要加入的任务")
            return

        # 检查是否有重复关键词
        duplicate_keywords = []
        for row in selected_rows:
            keyword = self.preview_table.item(row, 1).text()
            if database.check_keyword_exists(keyword):
                duplicate_keywords.append(keyword)

        # 如果有重复关键词，弹出确认对话框
        if duplicate_keywords:
            keywords_str = "\n".join([f"  - {kw}" for kw in duplicate_keywords])
            reply = QMessageBox.question(
                self, "重复关键词确认",
                f"以下关键词已在历史记录中：\n{keywords_str}\n\n是否仍要添加这些任务？",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                return

        added = 0
        skipped = 0
        for row in selected_rows:
            keyword = self.preview_table.item(row, 1).text()
            sort_type = self.preview_table.item(row, 2).text()
            count_str = self.preview_table.item(row, 3).text()
            channel_combo = self.preview_table.cellWidget(row, 4)
            channel = channel_combo.currentText()

            # 添加到数据库
            try:
                database.add_task(keyword, sort_type, int(count_str) if count_str.isdigit() else 5, channel)
                database.add_keyword_history(keyword)
                added += 1
            except Exception as e:
                skipped += 1

        self.statusBar.showMessage(f"添加了 {added} 个任务" + (f"，跳过 {skipped} 个" if skipped else ""))
        self.refresh_task_table()

    def add_all_to_tasks(self):
        self.preview_table.setRowCount(0)
        for data in self.preview_data:
            row = self.preview_table.rowCount()
            self.preview_table.insertRow(row)

            chk = QCheckBox()
            chk.setChecked(True)
            self.preview_table.setCellWidget(row, 0, chk)
            self.preview_table.setItem(row, 1, QTableWidgetItem(data["keyword"]))
            self.preview_table.setItem(row, 2, QTableWidgetItem(data["sort"]))
            self.preview_table.setItem(row, 3, QTableWidgetItem(data["count"]))

            channel_combo = QComboBox()
            channel_combo.addItems(["B站", "抖音", "youtube"])
            self.preview_table.setCellWidget(row, 4, channel_combo)

            if data.get("duplicate"):
                for col in range(1, 4):
                    item = self.preview_table.item(row, col)
                    if item:
                        item.setForeground(QColor("red"))

    def clear_completed_tasks(self):
        database.clear_completed_tasks()
        self.refresh_task_table()
        self.statusBar.showMessage("已清空已完成的任务")

    # ========== 刷新表格 ==========

    def refresh_task_table(self):
        tasks = database.get_all_tasks()
        self.task_table.setRowCount(len(tasks))
        for idx, task in enumerate(tasks):
            self.task_table.setItem(idx, 0, QTableWidgetItem(task["keyword"]))
            self.task_table.setItem(idx, 1, QTableWidgetItem(task["order_type"]))
            self.task_table.setItem(idx, 2, QTableWidgetItem(str(task["limit_count"])))
            self.task_table.setItem(idx, 3, QTableWidgetItem(task["channel"]))
            self.task_table.setItem(idx, 4, QTableWidgetItem(task["status"]))

            # 根据状态设置颜色
            status = task["status"]
            color = QColor("green") if status == "completed" else QColor("orange") if status == "pending" else QColor("gray")
            self.task_table.item(idx, 4).setForeground(color)

    def refresh_video_table(self):
        videos = database.get_all_videos()
        self.history_table.setRowCount(len(videos))
        for idx, video in enumerate(videos):
            self.history_table.setItem(idx, 0, QTableWidgetItem(video["title"]))
            self.history_table.setItem(idx, 1, QTableWidgetItem(video.get("keyword", "")))
            self.history_table.setItem(idx, 2, QTableWidgetItem(f"{video.get('file_size', 0):.1f}"))
            self.history_table.setItem(idx, 3, QTableWidgetItem(video.get("duration", "")))
            self.history_table.setItem(idx, 4, QTableWidgetItem(str(video.get("favorites", 0))))

            # 添加到ASR按钮
            btn = QPushButton("→ ASR")
            btn.clicked.connect(lambda _, v=video: self.add_single_video_to_asr(v))
            self.history_table.setCellWidget(idx, 5, btn)

    def refresh_asr_table(self):
        transcriptions = database.get_all_transcriptions()
        self.asr_table.setRowCount(len(transcriptions))
        for idx, t in enumerate(transcriptions):
            video_name = Path(t["video_path"]).stem if t["video_path"] else "未知"
            self.asr_table.setItem(idx, 0, QTableWidgetItem(video_name[:40]))
            self.asr_table.setItem(idx, 1, QTableWidgetItem(t["status"]))
            output_path = t.get("output_path") or ""
            self.asr_table.setItem(idx, 2, QTableWidgetItem(output_path[:40] if output_path else "无文件"))

            # 根据状态设置颜色
            status = t["status"]
            color = QColor("green") if status == "completed" else QColor("orange") if status == "pending" else QColor("red")
            self.asr_table.item(idx, 1).setForeground(color)

            # 提炼按钮
            btn = QPushButton("→ 提炼")
            btn.clicked.connect(lambda _, tid=t["id"], op=t.get("output_path") or "": self.add_to_summary(tid, op))
            self.asr_table.setCellWidget(idx, 3, btn)

        # 刷新转文字文件列表
        completed = [t for t in transcriptions if t["status"] == "completed"]
        self.transcription_table.setRowCount(len(completed))
        for idx, t in enumerate(completed):
            output_name = Path(t["output_path"]).name if t["output_path"] else "未知"
            video_name = Path(t["video_path"]).stem if t["video_path"] else "未知"
            self.transcription_table.setItem(idx, 0, QTableWidgetItem(output_name))
            self.transcription_table.setItem(idx, 1, QTableWidgetItem(video_name[:30]))
            self.transcription_table.setItem(idx, 2, QTableWidgetItem(t.get("completed_at", "")))

    def refresh_summary_table(self):
        summaries = database.get_all_summaries()
        self.summary_table.setRowCount(len(summaries))
        for idx, s in enumerate(summaries):
            output_name = Path(s["output_path"]).name if s["output_path"] else "处理中"
            input_name = Path(s["input_path"]).name if s["input_path"] else "未知"
            self.summary_table.setItem(idx, 0, QTableWidgetItem(output_name))
            self.summary_table.setItem(idx, 1, QTableWidgetItem(input_name[:30]))
            self.summary_table.setItem(idx, 2, QTableWidgetItem(s["status"]))

            status = s["status"]
            color = QColor("green") if status == "completed" else QColor("orange") if status == "pending" else QColor("red")
            self.summary_table.item(idx, 2).setForeground(color)

    def refresh_all_tables(self):
        self.refresh_task_table()
        self.refresh_video_table()
        self.refresh_asr_table()
        self.refresh_summary_table()

    # ========== 视频下载 ==========

    def start_crawling(self):
        tasks = database.get_pending_tasks()
        if not tasks:
            QMessageBox.information(self, "提示", "没有待执行的任务")
            return

        self.current_tasks = tasks
        self.start_crawl_btn.setEnabled(False)
        self.progress_bar.setMaximum(len(tasks))
        self.progress_bar.setValue(0)

        self.download_worker = DownloadWorker(tasks)
        self.download_worker.progress.connect(self.on_download_progress)
        self.download_worker.progress_value.connect(self.on_download_progress_value)
        self.download_worker.finished.connect(self.on_download_finished)
        self.download_worker.video_added.connect(self.on_video_added)
        self.download_worker.start()

    def on_download_progress(self, msg):
        self.statusBar.showMessage(msg)

    def on_download_progress_value(self, current, total):
        self.progress_bar.setValue(current)

    def on_download_finished(self, success, msg):
        self.start_crawl_btn.setEnabled(True)
        self.statusBar.showMessage(msg)
        self.refresh_task_table()
        self.refresh_video_table()

    def on_video_added(self, video):
        pass

    # ========== ASR 转文字 ==========

    def add_selected_videos_to_asr(self):
        """将选中的视频添加到ASR队列（去重）"""
        selected_rows = self.history_table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.information(self, "提示", "请先选择要转换的视频")
            return

        videos = database.get_all_videos()
        added = 0
        skipped = 0
        for row in selected_rows:
            row_idx = row.row()
            if row_idx < len(videos):
                video = videos[row_idx]
                video_id = video.get("id")
                video_path = video.get("file_path")
                
                if not video_path:
                    continue
                
                # 检查是否已存在
                if (video_id and database.is_video_in_transcriptions(video_id)) or \
                   database.is_video_path_in_transcriptions(video_path):
                    skipped += 1
                    continue
                
                database.add_transcription(video_id, video_path)
                added += 1

        msg_parts = []
        if added > 0:
            msg_parts.append(f"已添加 {added} 个视频到转文字队列")
        if skipped > 0:
            msg_parts.append(f"跳过 {skipped} 个重复项")
        
        if added > 0:
            QMessageBox.information(self, "成功", "\n".join(msg_parts))
            self.refresh_asr_table()
        elif skipped > 0:
            QMessageBox.information(self, "提示", f"选中的视频都已在队列中（{skipped} 个）")
        else:
            QMessageBox.warning(self, "警告", "选中的视频没有有效的文件路径")

    def add_single_video_to_asr(self, video):
        """添加单个视频到ASR（去重）"""
        if not video.get("file_path"):
            QMessageBox.warning(self, "警告", "视频文件路径无效")
            return
        
        # 检查是否已存在
        video_id = video.get("id")
        video_path = video["file_path"]
        
        if video_id and database.is_video_in_transcriptions(video_id):
            QMessageBox.information(self, "提示", f"「{video['title'][:20]}...」已经在转文字队列中了")
            return
        
        if database.is_video_path_in_transcriptions(video_path):
            QMessageBox.information(self, "提示", f"「{video['title'][:20]}...」已经在转文字队列中了")
            return
        
        database.add_transcription(video.get("id"), video_path)
        self.refresh_asr_table()
        self.statusBar.showMessage(f"已添加: {video['title'][:30]}")

    def start_transcription(self):
        tasks = database.get_pending_transcriptions()
        if not tasks:
            QMessageBox.information(self, "提示", "没有待处理的转文字任务")
            return

        self.start_transcribe_btn.setEnabled(False)
        self.statusBar.showMessage("开始转文字任务...")

        self.asr_worker = ASRWorker(tasks)
        self.asr_worker.progress.connect(self.statusBar.showMessage)
        self.asr_worker.finished.connect(self.on_transcription_finished)
        self.asr_worker.start()

    def on_transcription_finished(self, success, msg):
        self.start_transcribe_btn.setEnabled(True)
        self.statusBar.showMessage(msg)
        self.refresh_asr_table()

    # ========== 提炼总结 ==========

    def add_to_summary(self, transcription_id, output_path):
        """添加转文字结果到提炼队列"""
        if output_path and Path(output_path).exists():
            database.add_summary(transcription_id, output_path)
            self.refresh_summary_table()
            self.statusBar.showMessage("已添加到提炼队列")
        else:
            QMessageBox.warning(self, "警告", "转文字文件不存在")

    def generate_summary_from_trans(self):
        """从转文字文件生成总结"""
        # 获取所有已完成的转文字文件
        transcriptions = database.get_all_transcriptions()
        completed = [t for t in transcriptions if t["status"] == "completed" and t.get("output_path")]

        if not completed:
            QMessageBox.information(self, "提示", "没有已完成的转文字文件")
            return

        for t in completed:
            # 检查是否已经添加过
            summaries = database.get_all_summaries()
            already_added = any(s.get("transcription_id") == t["id"] for s in summaries)
            if not already_added:
                database.add_summary(t["id"], t["output_path"])

        self.refresh_summary_table()
        self.statusBar.showMessage(f"已将 {len(completed)} 个文件添加到提炼队列")

        # 开始处理
        tasks = database.get_pending_summaries()
        if tasks:
            self.generate_summary_btn.setEnabled(False)
            self.summary_worker = SummaryWorker(tasks)
            self.summary_worker.progress.connect(self.statusBar.showMessage)
            self.summary_worker.finished.connect(self.on_summary_finished)
            self.summary_worker.start()

    def on_summary_finished(self, success, msg):
        self.generate_summary_btn.setEnabled(True)
        self.statusBar.showMessage(msg)
        self.refresh_summary_table()

    # ========== 页面5: 设置 ==========
    def create_settings_page(self):
        from pathlib import Path
        BASE_DIR = Path(__file__).resolve().parent.parent
        
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setSpacing(16)

        # 数据库管理
        db_group = QGroupBox("数据库管理")
        db_layout = QVBoxLayout(db_group)

        # 当前状态
        status_layout = QHBoxLayout()
        status_layout.addWidget(QLabel("数据库文件:"))
        db_path = os.path.join(BASE_DIR, "sjcode.db")
        db_label = QLabel(db_path)
        db_label.setStyleSheet("color: #666; font-size: 11px;")
        status_layout.addWidget(db_label)
        status_layout.addStretch()
        db_layout.addLayout(status_layout)

        # 统计信息
        stats_group = QGroupBox("数据统计")
        stats_layout = QVBoxLayout(stats_group)

        stats = database.get_all_stats()
        for key, value in stats.items():
            row_layout = QHBoxLayout()
            row_layout.addWidget(QLabel(f"{key}:"))
            row_layout.addStretch()
            count_label = QLabel(str(value))
            count_label.setStyleSheet("font-weight: bold; color: #2c974b;")
            row_layout.addWidget(count_label)
            stats_layout.addLayout(row_layout)

        db_layout.addWidget(stats_group)

        # 操作按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        clear_videos_btn = QPushButton("清空视频列表")
        clear_videos_btn.setObjectName("dangerBtn")
        clear_videos_btn.clicked.connect(lambda: self.clear_table("videos", "视频列表"))
        btn_layout.addWidget(clear_videos_btn)

        clear_asr_btn = QPushButton("清空转文字队列")
        clear_asr_btn.setObjectName("dangerBtn")
        clear_asr_btn.clicked.connect(lambda: self.clear_table("transcriptions", "转文字队列"))
        btn_layout.addWidget(clear_asr_btn)

        clear_tasks_btn = QPushButton("清空任务列表")
        clear_tasks_btn.setObjectName("dangerBtn")
        clear_tasks_btn.clicked.connect(lambda: self.clear_table("tasks", "任务列表"))
        btn_layout.addWidget(clear_tasks_btn)

        clear_all_btn = QPushButton("清空所有数据")
        clear_all_btn.setObjectName("dangerBtn")
        clear_all_btn.clicked.connect(self.clear_all_data)
        btn_layout.addWidget(clear_all_btn)

        db_layout.addLayout(btn_layout)
        layout.addWidget(db_group)

        # 关于
        about_group = QGroupBox("关于")
        about_layout = QVBoxLayout(about_group)
        about_layout.addWidget(QLabel("SJCode - 视频内容处理工作台"))
        about_layout.addWidget(QLabel("版本: 1.0.0"))
        about_layout.addWidget(QLabel("功能: B站视频下载 / ASR转文字 / AI提炼总结"))
        layout.addWidget(about_group)

        layout.addStretch()
        return page

    def _refresh_settings_stats(self):
        """刷新设置页面的统计数据"""
        stats = database.get_stats()
        self.stats_label.setText(
            f"视频库: {stats['videos']} 条记录\n"
            f"转文字队列: {stats['transcriptions']} 条记录\n"
            f"总结记录: {stats['summaries']} 条记录\n"
            f"关键词历史: {stats['keyword_history']} 条记录"
        )

    def clear_table(self, table_name, display_name):
        reply = QMessageBox.question(
            self,
            "确认清空",
            f"确定要清空 {display_name} 吗？\n此操作不可恢复。",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            database.clear_table(table_name)
            self.refresh_all_tables()
            self._refresh_settings_stats()
            self.statusBar.showMessage(f"{display_name} 已清空")

    def clear_all_data(self):
        reply = QMessageBox.question(
            self,
            "警告：清空所有数据",
            "确定要清空所有数据吗？\n包括：任务、视频、转文字、提炼记录。\n此操作不可恢复！",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            tables = ['transcriptions', 'summaries', 'videos', 'tasks', 'keyword_history']
            for table in tables:
                database.clear_table(table)
            self.refresh_all_tables()
            self._refresh_settings_stats()
            self.statusBar.showMessage("所有数据已清空")

    def _refresh_all_tables_settings(self):
        """刷新所有表格（供设置页面使用）"""
        self.refresh_task_table()
        self.refresh_video_table()
        self.refresh_asr_table()
        self.refresh_summary_table()


# ===================== 程序入口 =====================

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = VideoProcessorApp()
    window.show()
    sys.exit(app.exec())
