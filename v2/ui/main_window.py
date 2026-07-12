# -*- coding: utf-8 -*-
"""
V2 主窗口
功能：
  - 上传Excel并解析
  - 视频下载管理
  - 视频转文本
  - Cookie管理
"""

import os
import sys
from pathlib import Path
from typing import List, Dict, Optional

from PySide6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
    QPushButton, QLabel, QTableWidget, QTableWidgetItem,
    QFileDialog, QMessageBox, QProgressBar, QTextEdit,
    QTabWidget, QHeaderView, QCheckBox, QGroupBox,
    QSplitter, QStatusBar
)
from PySide6.QtCore import Qt, QThread, Signal, QObject
from PySide6.QtGui import QFont

# 添加项目路径
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from v2.core.downloader import (
    download_video, download_videos_batch, 
    get_downloaded_videos, is_video_downloaded,
    save_cookies, DEFAULT_COOKIE_FILE
)
from v2.core.parser import (
    process_video, is_video_transcribed, get_transcription_status
)


# ===================== 工作线程 =====================

class DownloadWorker(QObject):
    """下载工作线程"""
    progress = Signal(str)
    finished = Signal(list)
    
    def __init__(self, videos: List[Dict], output_dir: Path = None, cookie_file: Path = None):
        super().__init__()
        self.videos = videos
        self.output_dir = output_dir
        self.cookie_file = cookie_file
    
    def run(self):
        results = download_videos_batch(
            self.videos, 
            self.output_dir, 
            self.cookie_file,
            lambda msg: self.progress.emit(msg)
        )
        self.finished.emit(results)


class TranscribeWorker(QObject):
    """转写工作线程"""
    progress = Signal(str)
    finished = Signal(dict)
    
    def __init__(self, video_path: str, title: str = ""):
        super().__init__()
        self.video_path = video_path
        self.title = title
    
    def run(self):
        result = process_video(
            self.video_path,
            self.title,
            progress_callback=lambda msg: self.progress.emit(msg)
        )
        self.finished.emit(result)


# ===================== 主窗口 =====================

class MainWindow(QMainWindow):
    """V2 主窗口"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("SJCode V2 - 视频处理工具")
        self.setMinimumSize(1000, 700)
        
        # 数据
        self.excel_data: List[Dict] = []
        self.current_excel_path: str = ""
        
        # 工作线程
        self.download_thread: Optional[QThread] = None
        self.download_worker: Optional[DownloadWorker] = None
        self.transcribe_thread: Optional[QThread] = None
        self.transcribe_worker: Optional[TranscribeWorker] = None
        
        self.init_ui()
    
    def init_ui(self):
        """初始化UI"""
        # 中央Widget
        central = QWidget()
        self.setCentralWidget(central)
        
        # 主布局
        main_layout = QVBoxLayout(central)
        
        # 创建TabWidget
        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs)
        
        # Tab 1: Excel管理
        self.create_excel_tab()
        
        # Tab 2: 视频库
        self.create_video_library_tab()
        
        # Tab 3: Cookie管理
        self.create_cookie_tab()
        
        # 状态栏
        self.statusBar().showMessage("就绪")
    
    # ===================== Excel Tab =====================
    
    def create_excel_tab(self):
        """创建Excel管理Tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # 顶部按钮区
        btn_layout = QHBoxLayout()
        
        self.upload_btn = QPushButton("上传Excel")
        self.upload_btn.clicked.connect(self.upload_excel)
        btn_layout.addWidget(self.upload_btn)
        
        self.download_selected_btn = QPushButton("下载选中视频")
        self.download_selected_btn.clicked.connect(self.download_selected)
        self.download_selected_btn.setEnabled(False)
        btn_layout.addWidget(self.download_selected_btn)
        
        self.download_all_btn = QPushButton("下载全部视频")
        self.download_all_btn.clicked.connect(self.download_all)
        self.download_all_btn.setEnabled(False)
        btn_layout.addWidget(self.download_all_btn)
        
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        # 表格
        self.excel_table = QTableWidget()
        self.excel_table.setColumnCount(7)
        self.excel_table.setHorizontalHeaderLabels([
            "选择", "BVID", "标题", "作者", "时长", "下载状态", "转写状态"
        ])
        self.excel_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.excel_table.setSelectionBehavior(QTableWidget.SelectRows)
        layout.addWidget(self.excel_table)
        
        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        # 日志区
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMaximumHeight(150)
        layout.addWidget(self.log_text)
        
        self.tabs.addTab(widget, "Excel管理")
    
    def upload_excel(self):
        """上传Excel文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择Excel文件", "", 
            "Excel文件 (*.xlsx *.xls);;所有文件 (*)"
        )
        
        if not file_path:
            return
        
        self.current_excel_path = file_path
        self.log(f"读取Excel: {file_path}")
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # 读取表头
            headers = [cell.value for cell in ws[1]]
            self.log(f"表头: {headers}")
            
            # 读取数据
            self.excel_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                data = dict(zip(headers, row))
                bvid = str(data.get("视频讯号(bvid)") or data.get("视频讯号") or data.get("bvid") or "").strip()
                if not bvid:
                    continue
                
                self.excel_data.append({
                    "bvid": bvid,
                    "title": str(data.get("视频标题") or data.get("title") or "").strip(),
                    "author": str(data.get("视频作者") or data.get("author") or "").strip(),
                    "duration": str(data.get("播放时长") or data.get("duration") or "").strip(),
                    "category": str(data.get("类别") or data.get("category") or "").strip(),
                    "keyword": str(data.get("检索关键词") or data.get("keyword") or "").strip(),
                    "url": f"https://www.bilibili.com/video/{bvid}",
                })
            
            wb.close()
            self.log(f"读取 {len(self.excel_data)} 条记录")
            
            # 更新表格
            self.refresh_excel_table()
            
            # 启用按钮
            self.download_selected_btn.setEnabled(True)
            self.download_all_btn.setEnabled(True)
            
        except Exception as e:
            self.log(f"读取Excel失败: {e}")
    
    def refresh_excel_table(self):
        """刷新Excel表格"""
        self.excel_table.setRowCount(len(self.excel_data))
        
        for i, video in enumerate(self.excel_data):
            bvid = video.get("bvid", "")
            
            # 选择框
            checkbox = QCheckBox()
            checkbox.setChecked(True)
            self.excel_table.setCellWidget(i, 0, checkbox)
            
            # BVID
            self.excel_table.setItem(i, 1, QTableWidgetItem(bvid))
            
            # 标题
            self.excel_table.setItem(i, 2, QTableWidgetItem(video.get("title", "")))
            
            # 作者
            self.excel_table.setItem(i, 3, QTableWidgetItem(video.get("author", "")))
            
            # 时长
            self.excel_table.setItem(i, 4, QTableWidgetItem(video.get("duration", "")))
            
            # 下载状态
            downloaded = is_video_downloaded(bvid)
            status_item = QTableWidgetItem("✓ 已下载" if downloaded else "未下载")
            status_item.setForeground(Qt.green if downloaded else Qt.gray)
            self.excel_table.setItem(i, 5, status_item)
            
            # 转写状态
            # 需要获取视频路径才能检查
            transcribed = "未知"
            status_item2 = QTableWidgetItem(transcribed)
            status_item2.setForeground(Qt.gray)
            self.excel_table.setItem(i, 6, status_item2)
    
    def get_selected_videos(self) -> List[Dict]:
        """获取选中的视频"""
        selected = []
        for i in range(self.excel_table.rowCount()):
            checkbox = self.excel_table.cellWidget(i, 0)
            if checkbox and checkbox.isChecked():
                selected.append(self.excel_data[i])
        return selected
    
    def download_selected(self):
        """下载选中的视频"""
        videos = self.get_selected_videos()
        if not videos:
            self.log("请先选择要下载的视频")
            return
        
        self.start_download(videos)
    
    def download_all(self):
        """下载全部视频"""
        if not self.excel_data:
            self.log("没有可下载的视频")
            return
        
        self.start_download(self.excel_data)
    
    def start_download(self, videos: List[Dict]):
        """开始下载"""
        self.log(f"\n开始下载 {len(videos)} 个视频...")
        
        # 创建线程
        self.download_thread = QThread()
        self.download_worker = DownloadWorker(videos)
        self.download_worker.moveToThread(self.download_thread)
        
        # 连接信号
        self.download_thread.started.connect(self.download_worker.run)
        self.download_worker.progress.connect(self.log)
        self.download_worker.finished.connect(self.on_download_finished)
        self.download_worker.finished.connect(self.download_thread.quit)
        
        # 显示进度条
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, len(videos))
        self.progress_bar.setValue(0)
        
        # 禁用按钮
        self.download_selected_btn.setEnabled(False)
        self.download_all_btn.setEnabled(False)
        
        # 启动线程
        self.download_thread.start()
    
    def on_download_finished(self, results: List[Dict]):
        """下载完成"""
        success = sum(1 for r in results if r["status"] == "success")
        skipped = sum(1 for r in results if r["status"] == "skipped")
        failed = sum(1 for r in results if r["status"] == "failed")
        
        self.log(f"\n下载完成: 成功 {success}, 跳过 {skipped}, 失败 {failed}")
        
        # 隐藏进度条
        self.progress_bar.setVisible(False)
        
        # 启用按钮
        self.download_selected_btn.setEnabled(True)
        self.download_all_btn.setEnabled(True)
        
        # 刷新表格
        self.refresh_excel_table()
        
        # 刷新视频库
        self.refresh_video_library()
        
        # 不弹窗，只在日志显示结果
    
    # ===================== 视频库 Tab =====================
    
    def create_video_library_tab(self):
        """创建视频库Tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # 顶部按钮区
        btn_layout = QHBoxLayout()
        
        self.refresh_btn = QPushButton("刷新列表")
        self.refresh_btn.clicked.connect(self.refresh_video_library)
        btn_layout.addWidget(self.refresh_btn)
        
        self.transcribe_selected_btn = QPushButton("转写选中视频")
        self.transcribe_selected_btn.clicked.connect(self.transcribe_selected)
        btn_layout.addWidget(self.transcribe_selected_btn)
        
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        # 表格
        self.video_table = QTableWidget()
        self.video_table.setColumnCount(5)
        self.video_table.setHorizontalHeaderLabels([
            "选择", "BVID", "标题", "文件大小", "转写状态"
        ])
        self.video_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.video_table.setSelectionBehavior(QTableWidget.SelectRows)
        layout.addWidget(self.video_table)
        
        # 日志区
        self.video_log = QTextEdit()
        self.video_log.setReadOnly(True)
        self.video_log.setMaximumHeight(150)
        layout.addWidget(self.video_log)
        
        self.tabs.addTab(widget, "视频库")
    
    def refresh_video_library(self):
        """刷新视频库"""
        videos = get_downloaded_videos()
        
        self.video_table.setRowCount(len(videos))
        
        for i, video in enumerate(videos):
            bvid = video.get("bvid", "")
            path = video.get("path", "")
            
            # 选择框
            checkbox = QCheckBox()
            self.video_table.setCellWidget(i, 0, checkbox)
            
            # BVID
            self.video_table.setItem(i, 1, QTableWidgetItem(bvid))
            
            # 标题
            self.video_table.setItem(i, 2, QTableWidgetItem(video.get("title", "")))
            
            # 文件大小
            size_mb = video.get("size", 0) / (1024 * 1024)
            self.video_table.setItem(i, 3, QTableWidgetItem(f"{size_mb:.1f} MB"))
            
            # 转写状态
            transcribed = is_video_transcribed(path)
            status = "✓ 已转写" if transcribed else "未转写"
            status_item = QTableWidgetItem(status)
            status_item.setForeground(Qt.green if transcribed else Qt.gray)
            self.video_table.setItem(i, 4, status_item)
    
    def transcribe_selected(self):
        """转写选中的视频"""
        selected = []
        for i in range(self.video_table.rowCount()):
            checkbox = self.video_table.cellWidget(i, 0)
            if checkbox and checkbox.isChecked():
                bvid = self.video_table.item(i, 1).text()
                title = self.video_table.item(i, 2).text()
                # 获取路径
                videos = get_downloaded_videos()
                for v in videos:
                    if v["bvid"] == bvid:
                        selected.append({"path": v["path"], "title": title})
                        break
        
        if not selected:
            self.log("请先选择要转写的视频")
            return
        
        # 开始转写第一个
        video = selected[0]
        self.start_transcribe(video["path"], video["title"])
    
    def start_transcribe(self, video_path: str, title: str = ""):
        """开始转写"""
        self.video_log.append(f"\n开始转写: {title}")
        
        # 创建线程
        self.transcribe_thread = QThread()
        self.transcribe_worker = TranscribeWorker(video_path, title)
        self.transcribe_worker.moveToThread(self.transcribe_thread)
        
        # 连接信号
        self.transcribe_thread.started.connect(self.transcribe_worker.run)
        self.transcribe_worker.progress.connect(lambda msg: self.video_log.append(msg))
        self.transcribe_worker.finished.connect(self.on_transcribe_finished)
        self.transcribe_worker.finished.connect(self.transcribe_thread.quit)
        
        # 禁用按钮
        self.transcribe_selected_btn.setEnabled(False)
        
        # 启动线程
        self.transcribe_thread.start()
    
    def on_transcribe_finished(self, result: Dict):
        """转写完成"""
        if result["status"] == "success":
            self.video_log.append(f"\n转写完成: {result['markdown_path']}")
        else:
            self.video_log.append(f"\n转写失败: {result['message']}")
        
        # 启用按钮
        self.transcribe_selected_btn.setEnabled(True)
        
        # 刷新列表
        self.refresh_video_library()
    
    # ===================== Cookie Tab =====================
    
    def create_cookie_tab(self):
        """创建Cookie管理Tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # 说明
        info_label = QLabel(
            "Cookie 用于下载B站视频。\n"
            "请从浏览器导出 cookies.txt 文件（Netscape 格式），然后上传。\n\n"
            "导出方法：\n"
            "1. 安装 Chrome 插件 'Get cookies.txt'\n"
            "2. 登录 bilibili.com\n"
            "3. 点击插件导出 cookies"
        )
        info_label.setWordWrap(True)
        layout.addWidget(info_label)
        
        # 按钮区
        btn_layout = QHBoxLayout()
        
        self.upload_cookie_btn = QPushButton("上传Cookie文件")
        self.upload_cookie_btn.clicked.connect(self.upload_cookie)
        btn_layout.addWidget(self.upload_cookie_btn)
        
        self.view_cookie_btn = QPushButton("查看当前Cookie")
        self.view_cookie_btn.clicked.connect(self.view_cookie)
        btn_layout.addWidget(self.view_cookie_btn)
        
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        # Cookie内容显示
        self.cookie_text = QTextEdit()
        self.cookie_text.setReadOnly(True)
        layout.addWidget(self.cookie_text)
        
        # 当前Cookie文件路径
        path_label = QLabel(f"Cookie文件: {DEFAULT_COOKIE_FILE}")
        layout.addWidget(path_label)
        
        self.tabs.addTab(widget, "Cookie管理")
    
    def upload_cookie(self):
        """上传Cookie文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择Cookie文件", "", 
            "文本文件 (*.txt);;所有文件 (*)"
        )
        
        if not file_path:
            return
        
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                content = f.read()
            
            # 保存到默认位置
            if save_cookies(content):
                self.log("Cookie上传成功!")
                self.cookie_text.setText(content)
            else:
                self.log("Cookie保存失败!")
        
        except Exception as e:
            self.log(f"上传失败: {e}")
    
    def view_cookie(self):
        """查看当前Cookie"""
        if DEFAULT_COOKIE_FILE.exists():
            try:
                with open(DEFAULT_COOKIE_FILE, "r", encoding="utf-8") as f:
                    content = f.read()
                self.cookie_text.setText(content)
            except Exception as e:
                self.cookie_text.setText(f"读取失败: {e}")
        else:
            self.cookie_text.setText("Cookie文件不存在")
    
    # ===================== 工具方法 =====================
    
    def log(self, message: str):
        """输出日志"""
        self.log_text.append(message)
        self.statusBar().showMessage(message)


# ===================== 入口 =====================

def main():
    """主入口"""
    from PySide6.QtWidgets import QApplication
    
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
